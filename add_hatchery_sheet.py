"""
FarmWell — Add Hatchery Audit Sheet to biosecurity_assessments_export.xlsx
Extracts questions/fields from Steps 2-8 of the hatchery audit JSX components.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# MANUALLY EXTRACTED HATCHERY AUDIT FIELDS
# Source: src/modules/poultry/components/hatchery/audit/steps/Step2-8.jsx
# ---------------------------------------------------------------------------

hatchery_questions = [

    # ── STEP 2: Vaccine Storage (Step2_VaccineStorage.jsx) ──────────────────
    {
        'Question ID': 'H1',
        'Order': 1,
        'Category': 'Vaccine Storage',
        'Question (EN)': 'Current Refrigerator Temperature (°C)',
        'Question (VI)': 'Nhiệt độ Tủ lạnh Hiện tại (°C)',
        'Answer Type': 'number_input',
        'Options': 'Valid range: 2–8°C (auto-validated)',
        'Risk Priority': '',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H2',
        'Order': 2,
        'Category': 'Vaccine Storage',
        'Question (EN)': 'Refrigerator temperature maintained at +2°C to +8°C',
        'Question (VI)': 'Nhiệt độ tủ lạnh duy trì ở +2°C đến +8°C',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No (checked = compliant)',
        'Risk Priority': 'high',
        'Related Diseases': 'All vaccine-preventable diseases',
    },
    {
        'Question ID': 'H3',
        'Order': 3,
        'Category': 'Vaccine Storage',
        'Question (EN)': 'Temperature monitoring log available and up-to-date',
        'Question (VI)': 'Nhật ký theo dõi nhiệt độ có sẵn và cập nhật',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'medium',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H4',
        'Order': 4,
        'Category': 'Vaccine Storage',
        'Question (EN)': 'Refrigerator used for vaccine storage only (no food/beverages)',
        'Question (VI)': 'Tủ lạnh chỉ dùng để bảo quản vắc-xin (không có thức ăn/đồ uống)',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H5',
        'Order': 5,
        'Category': 'Vaccine Storage',
        'Question (EN)': "Vaccines arranged by 'First to expire, first out' (FEFO) principle",
        'Question (VI)': "Vắc-xin được sắp xếp theo nguyên tắc 'Hết hạn trước, dùng trước'",
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'medium',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H6',
        'Order': 6,
        'Category': 'Vaccine Storage',
        'Question (EN)': 'Vaccines stored in proper position (not on door shelves)',
        'Question (VI)': 'Vắc-xin được lưu trữ ở vị trí đúng (không trên kệ cửa)',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'medium',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H7',
        'Order': 7,
        'Category': 'Vaccine Storage',
        'Question (EN)': 'Clear labeling for vaccines with cold chain incidents',
        'Question (VI)': 'Dán nhãn rõ ràng cho vắc-xin có sự cố chuỗi lạnh',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'medium',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H8',
        'Order': 8,
        'Category': 'Vaccine Storage',
        'Question (EN)': 'Notes & Observations',
        'Question (VI)': 'Ghi chú & Quan sát',
        'Answer Type': 'textarea',
        'Options': 'Free text — issues, observations, corrective actions',
        'Risk Priority': '',
        'Related Diseases': '',
    },

    # ── STEP 3: Equipment (Step3_Equipment.jsx) ──────────────────────────────
    {
        'Question ID': 'H9',
        'Order': 9,
        'Category': 'Vaccination Equipment',
        'Question (EN)': 'Equipment Type',
        'Question (VI)': 'Loại Thiết bị',
        'Answer Type': 'select',
        'Options': 'Spray Cabinet (Henke Sass Wolf); Pneumatic Vaccinator; Other',
        'Risk Priority': '',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H10',
        'Order': 10,
        'Category': 'Vaccination Equipment',
        'Question (EN)': 'Equipment Name / Model',
        'Question (VI)': 'Tên/Model Thiết bị',
        'Answer Type': 'text_input',
        'Options': 'e.g., Henke Sass Wolf Model X',
        'Risk Priority': '',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H11',
        'Order': 11,
        'Category': 'Vaccination Equipment',
        'Question (EN)': 'Quantity',
        'Question (VI)': 'Số lượng',
        'Answer Type': 'number_input',
        'Options': 'Min: 1',
        'Risk Priority': '',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H12',
        'Order': 12,
        'Category': 'Vaccination Equipment',
        'Question (EN)': 'Last Service Date',
        'Question (VI)': 'Ngày Bảo trì Cuối cùng',
        'Answer Type': 'date_input',
        'Options': '',
        'Risk Priority': 'medium',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H13',
        'Order': 13,
        'Category': 'Vaccination Equipment',
        'Question (EN)': 'Available Doses',
        'Question (VI)': 'Liều có sẵn',
        'Answer Type': 'number_input',
        'Options': 'Number of doses available',
        'Risk Priority': '',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H14',
        'Order': 14,
        'Category': 'Vaccination Equipment',
        'Question (EN)': 'Equipment in good working condition',
        'Question (VI)': 'Thiết bị hoạt động tốt',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H15',
        'Order': 15,
        'Category': 'Vaccination Equipment',
        'Question (EN)': 'Regular maintenance performed',
        'Question (VI)': 'Bảo trì định kỳ được thực hiện',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'medium',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H16',
        'Order': 16,
        'Category': 'Vaccination Equipment',
        'Question (EN)': 'Sufficient vaccine doses available',
        'Question (VI)': 'Có đủ liều vắc-xin',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H17',
        'Order': 17,
        'Category': 'Vaccination Equipment',
        'Question (EN)': 'Cleaning & disinfection protocol followed',
        'Question (VI)': 'Tuân thủ quy trình vệ sinh & khử trùng',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H18',
        'Order': 18,
        'Category': 'Vaccination Equipment',
        'Question (EN)': 'Spare parts inventory adequate',
        'Question (VI)': 'Kho phụ tùng đầy đủ',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'low',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H19',
        'Order': 19,
        'Category': 'Vaccination Equipment',
        'Question (EN)': 'Notes (per equipment item)',
        'Question (VI)': 'Ghi chú (mỗi thiết bị)',
        'Answer Type': 'textarea',
        'Options': 'Free text — any issues or observations',
        'Risk Priority': '',
        'Related Diseases': '',
    },

    # ── STEP 4: Techniques — Section A: Vaccine Preparation ────────────────
    {
        'Question ID': 'H20',
        'Order': 20,
        'Category': 'Vaccination Techniques – A. Preparation',
        'Question (EN)': 'Aseptic technique followed',
        'Question (VI)': 'Tuân thủ kỹ thuật vô trùng',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H21',
        'Order': 21,
        'Category': 'Vaccination Techniques – A. Preparation',
        'Question (EN)': 'Hand hygiene performed before preparation',
        'Question (VI)': 'Vệ sinh tay thực hiện trước khi chuẩn bị',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H22',
        'Order': 22,
        'Category': 'Vaccination Techniques – A. Preparation',
        'Question (EN)': 'New syringe used for each batch',
        'Question (VI)': 'Sử dụng kim tiêm mới cho mỗi lô',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H23',
        'Order': 23,
        'Category': 'Vaccination Techniques – A. Preparation',
        'Question (EN)': '21G needle used',
        'Question (VI)': 'Sử dụng kim 21G',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'medium',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H24',
        'Order': 24,
        'Category': 'Vaccination Techniques – A. Preparation',
        'Question (EN)': 'Expiration dates checked',
        'Question (VI)': 'Kiểm tra ngày hết hạn',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H25',
        'Order': 25,
        'Category': 'Vaccination Techniques – A. Preparation',
        'Question (EN)': 'Mixed according to manufacturer guidelines',
        'Question (VI)': 'Pha theo hướng dẫn của nhà sản xuất',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H26',
        'Order': 26,
        'Category': 'Vaccination Techniques – A. Preparation',
        'Question (EN)': 'Records maintained (date, time, batch, lot)',
        'Question (VI)': 'Hồ sơ được duy trì (ngày, giờ, lô, số lô)',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No',
        'Risk Priority': 'medium',
        'Related Diseases': '',
    },

    # ── STEP 4: Techniques — Section B: Spray Vaccination Quality ───────────
    {
        'Question ID': 'H27',
        'Order': 27,
        'Category': 'Vaccination Techniques – B. Spray Quality',
        'Question (EN)': 'Sample Size (number of trays observed)',
        'Question (VI)': 'Cỡ mẫu (số khay quan sát)',
        'Answer Type': 'number_input',
        'Options': 'Min: 0',
        'Risk Priority': '',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H28',
        'Order': 28,
        'Category': 'Vaccination Techniques – B. Spray Quality',
        'Question (EN)': 'Droplet Uniformity',
        'Question (VI)': 'Độ đồng đều của giọt',
        'Answer Type': 'select',
        'Options': 'Excellent; Good; Fair; Poor',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H29',
        'Order': 29,
        'Category': 'Vaccination Techniques – B. Spray Quality',
        'Question (EN)': 'Tray Coverage (%)',
        'Question (VI)': 'Độ bao phủ khay (%)',
        'Answer Type': 'number_input',
        'Options': 'Range: 0–100. Target: ≥90%',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },

    # ── STEP 4: Techniques — Section C: Injection Quality ───────────────────
    {
        'Question ID': 'H30',
        'Order': 30,
        'Category': 'Vaccination Techniques – C. Injection Quality',
        'Question (EN)': 'Sample Size (number of chicks observed)',
        'Question (VI)': 'Cỡ mẫu (số gà con quan sát)',
        'Answer Type': 'number_input',
        'Options': 'Min: 0',
        'Risk Priority': '',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H31',
        'Order': 31,
        'Category': 'Vaccination Techniques – C. Injection Quality',
        'Question (EN)': 'Accurate Injection (%)',
        'Question (VI)': 'Tiêm chính xác (%)',
        'Answer Type': 'number_input',
        'Options': 'Range: 0–100. Target: ≥95%',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H32',
        'Order': 32,
        'Category': 'Vaccination Techniques – C. Injection Quality',
        'Question (EN)': 'Bleeding (%)',
        'Question (VI)': 'Chảy máu (%)',
        'Answer Type': 'number_input',
        'Options': 'Range: 0–100. Alert if >5%',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H33',
        'Order': 33,
        'Category': 'Vaccination Techniques – C. Injection Quality',
        'Question (EN)': 'Wet Neck (%)',
        'Question (VI)': 'Cổ ướt (%)',
        'Answer Type': 'number_input',
        'Options': 'Range: 0–100. Alert if >5%',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H34',
        'Order': 34,
        'Category': 'Vaccination Techniques – C. Injection Quality',
        'Question (EN)': 'No Vaccine (%)',
        'Question (VI)': 'Không có vắc-xin (%)',
        'Answer Type': 'number_input',
        'Options': 'Range: 0–100. Alert if >5%',
        'Risk Priority': 'high',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H35',
        'Order': 35,
        'Category': 'Vaccination Techniques – C. Injection Quality',
        'Question (EN)': 'Notes & Observations',
        'Question (VI)': 'Ghi chú & Quan sát',
        'Answer Type': 'textarea',
        'Options': 'Free text — issues, observations, training needs',
        'Risk Priority': '',
        'Related Diseases': '',
    },

    # ── STEP 5: Sample Plan (Step5_SamplePlan.jsx) ───────────────────────────
    {
        'Question ID': 'H36',
        'Order': 36,
        'Category': 'Environmental Sampling Plan',
        'Question (EN)': 'Review and confirm 30-point environmental sampling plan',
        'Question (VI)': 'Xem lại và xác nhận kế hoạch lấy mẫu môi trường 30 điểm',
        'Answer Type': 'review_confirm',
        'Options': 'Standard protocol: 30 monitoring points across hatchery (eggshell, fluff, air, water, surface, chick samples)',
        'Risk Priority': 'high',
        'Related Diseases': 'Salmonella, Aspergillus, environmental pathogens',
    },

    # ── STEP 6: Sample Collection (Step6_SampleCollection.jsx) ──────────────
    {
        'Question ID': 'H37',
        'Order': 37,
        'Category': 'Sample Collection',
        'Question (EN)': 'Sample collected (per location point)',
        'Question (VI)': 'Mẫu đã thu thập (mỗi điểm lấy mẫu)',
        'Answer Type': 'checkbox',
        'Options': 'Yes / No (with timestamp auto-recorded on check)',
        'Risk Priority': 'high',
        'Related Diseases': 'Salmonella, Aspergillus, environmental pathogens',
    },
    {
        'Question ID': 'H38',
        'Order': 38,
        'Category': 'Sample Collection',
        'Question (EN)': 'Location-specific note (per sample point)',
        'Question (VI)': 'Ghi chú vị trí cụ thể (mỗi điểm lấy mẫu)',
        'Answer Type': 'text_input',
        'Options': 'Free text — specific location observation (visible after marking collected)',
        'Risk Priority': '',
        'Related Diseases': '',
    },

    # ── STEP 7: Incubation Tracking (Step7_Incubation.jsx) ──────────────────
    {
        'Question ID': 'H39',
        'Order': 39,
        'Category': 'Incubation Tracking',
        'Question (EN)': 'Start incubation timer (confirm all samples placed at 37°C)',
        'Question (VI)': 'Bắt đầu hẹn giờ ấp (xác nhận tất cả mẫu được đặt ở 37°C)',
        'Answer Type': 'action_button',
        'Options': 'Triggers incubation start — records timestamp automatically',
        'Risk Priority': '',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H40',
        'Order': 40,
        'Category': 'Incubation Tracking',
        'Question (EN)': 'Incubation Start Date & Time',
        'Question (VI)': 'Ngày & Giờ Bắt đầu Ấp',
        'Answer Type': 'datetime_input',
        'Options': 'Auto-set on timer start; editable',
        'Risk Priority': '',
        'Related Diseases': '',
    },
    {
        'Question ID': 'H41',
        'Order': 41,
        'Category': 'Incubation Tracking',
        'Question (EN)': 'Expected Completion Date & Time (72 hours)',
        'Question (VI)': 'Ngày & Giờ Dự kiến Hoàn thành (72 giờ)',
        'Answer Type': 'datetime_input',
        'Options': 'Auto-calculated as Start + 72 hours; editable',
        'Risk Priority': '',
        'Related Diseases': '',
    },

    # ── STEP 8: Results Entry (Step8_Results.jsx) ────────────────────────────
    {
        'Question ID': 'H42',
        'Order': 42,
        'Category': 'Results Entry',
        'Question (EN)': 'Aspergillus colony count (air plate samples only)',
        'Question (VI)': 'Số lượng khuẩn lạc Aspergillus (chỉ mẫu khay không khí)',
        'Answer Type': 'number_input',
        'Options': 'Min: 0 — per sample point; used in plate scoring algorithm',
        'Risk Priority': 'high',
        'Related Diseases': 'Aspergillosis',
    },
    {
        'Question ID': 'H43',
        'Order': 43,
        'Category': 'Results Entry',
        'Question (EN)': 'Other Molds / Total Colony Count (per sample point)',
        'Question (VI)': 'Nấm mốc khác / Tổng số khuẩn lạc (mỗi điểm lấy mẫu)',
        'Answer Type': 'number_input',
        'Options': 'Min: 0 — all sample types; used in scoring algorithm',
        'Risk Priority': 'high',
        'Related Diseases': 'Aspergillosis, fungal contamination',
    },
    {
        'Question ID': 'H44',
        'Order': 44,
        'Category': 'Results Entry',
        'Question (EN)': 'Individual sample score (1–5, auto-calculated)',
        'Question (VI)': 'Điểm mẫu đơn lẻ (1–5, tính tự động)',
        'Answer Type': 'calculated',
        'Options': '1=Excellent, 2=Good, 3=Fair, 4=Warning, 5=Critical — calculated from colony counts',
        'Risk Priority': 'high',
        'Related Diseases': 'All environmental pathogens',
    },
    {
        'Question ID': 'H45',
        'Order': 45,
        'Category': 'Results Entry',
        'Question (EN)': 'Overall Environmental Score (auto-calculated from all samples)',
        'Question (VI)': 'Điểm Môi trường Tổng thể (tính tự động từ tất cả mẫu)',
        'Answer Type': 'calculated',
        'Options': 'Composite score from all 30 sample points using weighted scoring engine',
        'Risk Priority': 'high',
        'Related Diseases': 'All environmental pathogens',
    },
]

# ---------------------------------------------------------------------------
# ADD HATCHERY SHEET TO EXISTING EXCEL
# ---------------------------------------------------------------------------

def add_hatchery_sheet():
    excel_file = os.path.join(BASE, 'biosecurity_assessments_export.xlsx')
    out_file   = os.path.join(BASE, 'biosecurity_assessments_v2.xlsx')

    if not os.path.exists(excel_file):
        print(f"ERROR: {excel_file} not found. Run generate_exports.py first.")
        return

    df_hatchery = pd.DataFrame(hatchery_questions)

    # Load all existing sheets into DataFrames
    existing = pd.read_excel(excel_file, sheet_name=None)  # dict of {name: df}

    # Write all sheets + new Hatchery sheet in one pass to a new file
    with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        for sheet_name, df in existing.items():
            if sheet_name != 'Hatchery':  # skip if old Hatchery exists
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        df_hatchery.to_excel(writer, sheet_name='Hatchery', index=False)

    # Apply formatting
    wb = load_workbook(out_file)
    ws = wb['Hatchery']

    fill_hex = '70AD47'
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
    font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=False, vertical='center')

    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((min(len(str(c.value)), 200) for c in col if c.value), default=10)
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

    wb.save(out_file)

    print(f"\n✅  biosecurity_assessments_v2.xlsx created (4 sheets)")
    print(f"    Total fields extracted: {len(df_hatchery)}")
    print(f"    Steps covered: 2–8 (Vaccine Storage, Equipment, Techniques A/B/C,")
    print(f"                         Sample Plan, Sample Collection, Incubation, Results)")
    print(f"\n    Sheet breakdown:")
    for cat, grp in df_hatchery.groupby('Category', sort=False):
        print(f"      {cat}: {len(grp)} fields")


if __name__ == '__main__':
    add_hatchery_sheet()
