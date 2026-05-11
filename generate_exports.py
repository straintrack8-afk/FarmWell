"""
FarmWell Data Export Generator
Generates Excel exports for:
  1. poultry_diseases_export.xlsx   (129 diseases, EN/VI)
  2. biosecurity_assessments_export.xlsx  (Broiler/Breeder/Layer, EN/VI)
"""

import json
import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def load_json(rel_path):
    full = os.path.join(BASE, rel_path)
    with open(full, 'r', encoding='utf-8') as f:
        return json.load(f)


def style_header(ws, fill_hex):
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
    font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=False, vertical='center')


def auto_width(ws, cap=60):
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value:
                # measure only first 200 chars to keep it fast
                max_len = max(max_len, min(len(str(cell.value)), 200))
        ws.column_dimensions[letter].width = min(max_len + 2, cap)


# ---------------------------------------------------------------------------
# PART 1 — DISEASE DATABASE
# ---------------------------------------------------------------------------

def build_disease_export():
    print("Loading disease JSON files...")
    data_en = load_json('public/data/poultry/diseases_COMPLETE_129_v4.1_ENRICHED_en.json')
    data_vi = load_json('public/data/poultry/diseases_COMPLETE_129_v4.1_ENRICHED_vi.json')

    # Both files wrap diseases under a 'diseases' key
    diseases_en = data_en['diseases']
    diseases_vi = {d['id']: d for d in data_vi['diseases']}

    rows = []
    for d_en in diseases_en:
        did = d_en['id']
        d_vi = diseases_vi.get(did, {})

        rows.append({
            'Disease ID':                    did,
            'Disease Name (EN)':             d_en.get('name', ''),
            'Disease Name (VI)':             d_vi.get('name', ''),
            'Category':                      d_en.get('category', ''),
            'Severity':                      d_en.get('severity', ''),
            'Zoonotic':                      'Yes' if d_en.get('zoonotic') else 'No',
            'Age Groups':                    ', '.join(d_en.get('ageGroups', [])),
            'Description (EN)':              d_en.get('description', ''),
            'Description (VI)':              d_vi.get('description', ''),
            'Clinical Signs (EN)':           d_en.get('clinicalSigns', ''),
            'Clinical Signs (VI)':           d_vi.get('clinicalSigns', ''),
            'Transmission (EN)':             d_en.get('transmission', ''),
            'Transmission (VI)':             d_vi.get('transmission', ''),
            'Diagnosis (EN)':                d_en.get('diagnosis', ''),
            'Diagnosis (VI)':                d_vi.get('diagnosis', ''),
            'Treatment (EN)':                d_en.get('treatment', ''),
            'Treatment (VI)':                d_vi.get('treatment', ''),
            'Control (EN)':                  d_en.get('control', ''),
            'Control (VI)':                  d_vi.get('control', ''),
            'Vaccine Recommendations (EN)':  d_en.get('vaccineRecommendations', ''),
            'Vaccine Recommendations (VI)':  d_vi.get('vaccineRecommendations', ''),
        })

    df = pd.DataFrame(rows).sort_values('Disease ID').reset_index(drop=True)

    out = os.path.join(BASE, 'poultry_diseases_export.xlsx')
    df.to_excel(out, index=False, sheet_name='Diseases')

    wb = load_workbook(out)
    ws = wb['Diseases']
    style_header(ws, '4472C4')  # blue
    auto_width(ws)
    wb.save(out)

    print(f"  ✅  poultry_diseases_export.xlsx — {len(df)} diseases, 21 columns")
    return len(df)


# ---------------------------------------------------------------------------
# PART 2 — BIOSECURITY ASSESSMENTS
# ---------------------------------------------------------------------------

def format_options(options):
    """Format options list → 'Label EN (score: X); ...'"""
    if not options:
        return ''
    parts = []
    for opt in options:
        label = opt.get('label', {})
        label_en = label.get('en', '') if isinstance(label, dict) else str(label)
        score = opt.get('score', '')
        parts.append(f"{label_en} (score: {score})" if score != '' else label_en)
    return '; '.join(parts)


def extract_questions_list(categories_list):
    """For broiler / layer — categories is a JSON array."""
    rows = []
    for cat in categories_list:
        cat_name_en = cat.get('name', {}).get('en', '') if isinstance(cat.get('name'), dict) else cat.get('name', '')
        for q in cat.get('questions', []):
            ra = q.get('risk_assessment', {}) or {}
            rows.append({
                'Question ID':      q.get('id', ''),
                'Order':            q.get('order', ''),
                'Category':         cat_name_en,
                'Question (EN)':    (q.get('question') or {}).get('en', ''),
                'Question (VI)':    (q.get('question') or {}).get('vi', ''),
                'Answer Type':      q.get('answer_type', ''),
                'Options':          format_options(q.get('options', [])),
                'Risk Priority':    ra.get('priority', ''),
                'Related Diseases': ', '.join(ra.get('diseases_affected', [])),
            })
    return pd.DataFrame(rows)


def extract_questions_dict(categories_dict):
    """For breeder — categories is a JSON object keyed A, B, C..."""
    rows = []
    for cat_id in sorted(categories_dict.keys()):
        cat = categories_dict[cat_id]
        cat_name_raw = cat.get('name', {})
        cat_name_en = cat_name_raw.get('en', cat_id) if isinstance(cat_name_raw, dict) else str(cat_name_raw)
        for q in cat.get('questions', []):
            ra = q.get('risk_assessment', {}) or {}
            rows.append({
                'Question ID':      q.get('id', ''),
                'Order':            q.get('order', ''),
                'Category':         f"{cat_id} – {cat_name_en}",
                'Question (EN)':    (q.get('question') or {}).get('en', ''),
                'Question (VI)':    (q.get('question') or {}).get('vi', ''),
                'Answer Type':      q.get('answer_type', ''),
                'Options':          format_options(q.get('options', [])),
                'Risk Priority':    ra.get('priority', ''),
                'Related Diseases': ', '.join(ra.get('diseases_affected', [])),
            })
    return pd.DataFrame(rows)


def build_biosecurity_export():
    print("Loading biosecurity JSON files...")
    broiler_data = load_json('public/data/poultry/broiler_assessment.json')
    breeder_data = load_json('public/data/poultry/breeder_assessment.json')
    layer_data   = load_json('public/data/poultry/layer_assessment_complete.json')

    # Broiler & Layer: categories is a list
    df_broiler = extract_questions_list(broiler_data.get('categories', []))
    df_layer   = extract_questions_list(layer_data.get('categories', []))

    # Breeder: categories is a dict {A: {...}, B: {...}, ...}
    breeder_cats = breeder_data.get('categories', {})
    if isinstance(breeder_cats, dict):
        df_breeder = extract_questions_dict(breeder_cats)
    else:
        df_breeder = extract_questions_list(breeder_cats)

    out = os.path.join(BASE, 'biosecurity_assessments_export.xlsx')

    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df_broiler.to_excel(writer, sheet_name='Broiler', index=False)
        df_breeder.to_excel(writer, sheet_name='Breeder', index=False)
        df_layer.to_excel(writer, sheet_name='Layer',   index=False)

    wb = load_workbook(out)
    fill_hex = '70AD47'  # green
    for sheet_name in ['Broiler', 'Breeder', 'Layer']:
        ws = wb[sheet_name]
        style_header(ws, fill_hex)
        auto_width(ws)
    wb.save(out)

    print(f"  ✅  biosecurity_assessments_export.xlsx")
    print(f"       Broiler: {len(df_broiler)} questions")
    print(f"       Breeder: {len(df_breeder)} questions")
    print(f"       Layer:   {len(df_layer)} questions")
    return len(df_broiler), len(df_breeder), len(df_layer)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    try:
        import openpyxl
    except ImportError:
        print("Missing dependency. Run:  pip install pandas openpyxl")
        sys.exit(1)

    print("\n=== FarmWell Excel Export ===\n")
    n_diseases = build_disease_export()
    b, br, la = build_biosecurity_export()

    print("\n=== Summary ===")
    print(f"  poultry_diseases_export.xlsx       {n_diseases} rows")
    print(f"  biosecurity_assessments_export.xlsx  Broiler={b} | Breeder={br} | Layer={la}")
    print("\nDone! Files saved to project root.\n")
